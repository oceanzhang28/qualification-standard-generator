#!/usr/bin/env python3
"""将任职资格 Excel 案例转换为 Skill 可读取的 Markdown 案例库。"""

from __future__ import annotations

import re
from collections import Counter
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
SKILL_ROOT = ROOT / "qualification-standard-generator"
SOURCE_DIR = ROOT / "任职资格参考标准"
CASE_LIBRARY = SKILL_ROOT / "references" / "case_library"
FULL_POSITIONS = CASE_LIBRARY / "full_positions"
SNIPPETS = CASE_LIBRARY / "snippets"
BY_MODULE = SNIPPETS / "by_module"
BY_BEHAVIOR_ITEM = SNIPPETS / "by_behavior_item"
BY_LEVEL_PROGRESSION = SNIPPETS / "by_level_progression"
POSITION_MAP_OUTPUT = SKILL_ROOT / "references" / "position_map.md"
REPORT_OUTPUT = SKILL_ROOT / "references" / "case_library" / "conversion_report.md"

HEADERS = [
    "序号",
    "行为模块",
    "行为要项",
    "专员行为标准",
    "主管行为标准",
    "经理行为标准",
    "总监行为标准",
]

LEVEL_HEADERS = HEADERS[3:]


@dataclass(frozen=True)
class CaseMeta:
    case_id: int
    source_file: str
    job_class: str
    job_subclass: str
    department: str
    position_name: str
    levels: str
    maturity: str
    learn_from: str
    do_not_learn: str
    business_context: str
    transfer_reference: str


CASES = [
    CaseMeta(
        case_id=1,
        source_file="HRBP.xlsx",
        job_class="人力资源类",
        job_subclass="HRBP子类",
        department="人力资源管理中心",
        position_name="HRBP岗",
        levels="专员、主管、经理、总监",
        maturity="成熟",
        learn_from="模块拆分清晰；战略规划、业务赋能、日常运作等模块贴合 HRBP 工作；职级差异明显；行为要项颗粒度合适。",
        do_not_learn="无",
        business_context="无",
        transfer_reference="职能岗位可参考",
    ),
    CaseMeta(
        case_id=2,
        source_file="有纺产品企划.xlsx",
        job_class="产品类",
        job_subclass="产品企划子类",
        department="商品中心",
        position_name="有纺产品企划岗",
        levels="专员、主管、经理、总监",
        maturity="成熟",
        learn_from="模块拆分清晰，采用流程化方式覆盖完整工作流程。",
        do_not_learn="部分职级间虽有工作细项差异，但具体工作内容存在复制情况。",
        business_context="负责有纺类型产品企划，公司另有无纺类型产品的全链条岗位配置。",
        transfer_reference="无纺产品企划、产品类岗位",
    ),
    CaseMeta(
        case_id=3,
        source_file="产品研发.xlsx",
        job_class="产品类",
        job_subclass="产品研发子类",
        department="研发部",
        position_name="产品研发岗",
        levels="专员、主管、经理、总监",
        maturity="成熟",
        learn_from="模块拆分清晰，兼具流程化和活动化；行为要项分解充分；职级差异通过动作体现；消费者洞察、基材研究、产品开发等要项颗粒度合适。",
        do_not_learn="暂无",
        business_context="负责无纺产品研发，公司另有有纺类型产品研发。",
        transfer_reference="产品类岗位",
    ),
    CaseMeta(
        case_id=4,
        source_file="产品内容策划.xlsx",
        job_class="产品类",
        job_subclass="产品营销子类",
        department="商品中心",
        position_name="产品内容策划岗",
        levels="专员、主管、经理",
        maturity="行为模块与要项成熟，行为标准不成熟",
        learn_from="模块拆分清晰；职级差异明显。",
        do_not_learn="行为标准颗粒度不足，仅简单描述具体动作与工作产出，只满足基本要求。",
        business_context="负责产品卖点详情内容策划，公司另有品牌内容和渠道内容岗位。",
        transfer_reference="内容运营子类岗位",
    ),
    CaseMeta(
        case_id=5,
        source_file="电商运营.xlsx",
        job_class="销售类",
        job_subclass="线上平台销售子类",
        department="电商事业部",
        position_name="电商运营岗",
        levels="专员、主管、经理",
        maturity="行为模块与要项成熟，行为标准不成熟",
        learn_from="模块拆分清晰；职级差异明显。",
        do_not_learn="部分行为标准颗粒度不足，仅简单描述具体动作与工作产出，只满足基本要求。",
        business_context="负责电商平台运营，岗位设置为线上销售类；公司另有线下门店、渠道 KA、大客户销售、直播运营、新零售运营等岗位。",
        transfer_reference="线上销售子类岗位、平台运营类岗位",
    ),
    CaseMeta(
        case_id=6,
        source_file="产品设计.xlsx",
        job_class="产品类",
        job_subclass="产品设计子类",
        department="商品中心",
        position_name="产品设计岗",
        levels="专员、主管、经理、总监",
        maturity="不完全成熟，可对照学习",
        learn_from="可参考其产品开发、产品管理、项目管理、组织贡献的精简模块结构，以及产品设计类工作细项拆分。",
        do_not_learn="行为标准存在颗粒度、句式和职级差异不稳定的问题，不应直接作为成熟优质写法套用。",
        business_context="负责有纺服装类产品设计，与产品企划、工艺版师等协作完成设计方案、样品和订货方案确认。",
        transfer_reference="产品设计子类岗位、产品类岗位可对照参考",
    ),
]


def text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip().replace("\r\n", "\n").replace("\r", "\n")


def one_line(value: str) -> str:
    return re.sub(r"\s+", " ", value).strip()


def filename_part(value: str) -> str:
    cleaned = re.sub(r'[\\/:*?"<>|#\[\]\n\r\t]+', "_", value).strip(" _")
    return cleaned or "未命名"


def yaml_scalar(value: str) -> str:
    escaped = value.replace("\\", "\\\\").replace('"', '\\"')
    return f'"{escaped}"'


def md_cell(value: str) -> str:
    return value.replace("|", "\\|").replace("\n", "<br>")


def find_standard_sheet(workbook):
    for worksheet in workbook.worksheets:
        max_row = min(worksheet.max_row, 12)
        for row_idx in range(1, max_row + 1):
            values = [
                text(worksheet.cell(row=row_idx, column=col_idx).value)
                for col_idx in range(1, 8)
            ]
            if values == HEADERS:
                return worksheet, row_idx
    raise ValueError("未找到 A-G 标准表头：序号/行为模块/行为要项/专员/主管/经理/总监行为标准")


def read_case_rows(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet, header_row = find_standard_sheet(workbook)
    rows = []
    for excel_row in worksheet.iter_rows(
        min_row=header_row + 1,
        max_row=worksheet.max_row,
        min_col=1,
        max_col=7,
        values_only=True,
    ):
        values = [text(value) for value in excel_row]
        if not any(values):
            continue
        row = dict(zip(HEADERS, values))
        row["原始行为要项"] = row["行为要项"]
        # 统一规则：组织贡献模块固定为“平台/流程建设、人员培养”。
        # 只规范行为要项名称，不改写各职级行为标准原文。
        if row["行为模块"] == "组织贡献" and row["行为要项"] == "人才培养":
            row["行为要项"] = "人员培养"
        rows.append(row)
    return worksheet.title, header_row, rows


def frontmatter(meta: CaseMeta, extra: dict[str, str] | None = None) -> str:
    values = {
        "case_id": str(meta.case_id),
        "职族": "专业族",
        "职类": meta.job_class,
        "职务子类": meta.job_subclass,
        "职位类别": "不适用",
        "所在部门": meta.department,
        "岗位名称": meta.position_name,
        "适用职级范围": meta.levels,
        "成熟度": meta.maturity,
        "用途": "结构参考、写法参考、职级递进参考",
        "禁止": "不得直接复制内容到目标岗位；不得学习标记为不成熟的低颗粒度行为标准",
        "来源文件": str(SOURCE_DIR / meta.source_file),
    }
    if extra:
        values.update(extra)
    lines = ["---"]
    for key, value in values.items():
        lines.append(f"{key}: {yaml_scalar(value)}")
    lines.append("---")
    return "\n".join(lines)


def module_groups(rows: list[dict[str, str]]):
    groups: dict[str, list[dict[str, str]]] = {}
    for row in rows:
        module = row["行为模块"] or "未命名模块"
        groups.setdefault(module, []).append(row)
    return groups


def write_full_position(meta: CaseMeta, sheet_name: str, header_row: int, rows: list[dict[str, str]]):
    output = FULL_POSITIONS / f"{filename_part(meta.job_class)}_{filename_part(meta.job_subclass)}_{filename_part(meta.position_name)}.md"
    modules = module_groups(rows)
    lines = [
        frontmatter(meta, {"案例类型": "完整岗位案例", "源Sheet": sheet_name, "表头行": str(header_row)}),
        "",
        f"# 岗位：{meta.position_name}",
        "",
        "## 一、案例评价",
        "",
        f"- 成熟度：{meta.maturity}",
        f"- 最值得学习的地方：{meta.learn_from}",
        f"- 不建议学习的地方：{meta.do_not_learn}",
        f"- 特殊业务背景：{meta.business_context}",
        f"- 适合迁移参考：{meta.transfer_reference}",
        "",
        "## 二、行为模块总览",
        "",
        "| 序号 | 行为模块 | 行为要项数量 |",
        "|---|---|---:|",
    ]
    for idx, (module, items) in enumerate(modules.items(), 1):
        lines.append(f"| {idx} | {md_cell(module)} | {len(items)} |")
    lines.extend(["", "## 三、完整任职资格标准", ""])
    for module_idx, (module, items) in enumerate(modules.items(), 1):
        lines.extend([f"### 行为模块 {module_idx}：{module}", ""])
        for item_idx, row in enumerate(items, 1):
            lines.extend(
                [
                    f"#### 行为要项 {item_idx}：{row['行为要项']}",
                    "",
                    f"- 原始序号：{row['序号']}",
                    "",
                    "| 职级 | 行为标准 |",
                    "|---|---|",
                ]
            )
            for level in LEVEL_HEADERS:
                level_name = level.replace("行为标准", "")
                lines.append(f"| {level_name} | {md_cell(row[level])} |")
            lines.append("")
    output.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")
    return output


def write_module_snippets(meta: CaseMeta, rows: list[dict[str, str]]):
    outputs = []
    for module_idx, (module, items) in enumerate(module_groups(rows).items(), 1):
        output = BY_MODULE / f"{meta.case_id:02d}_{filename_part(meta.position_name)}_{module_idx:02d}_{filename_part(module)}.md"
        lines = [
            frontmatter(meta, {"案例类型": "行为模块片段", "行为模块": module}),
            "",
            f"# 行为模块片段：{module}",
            "",
            f"- 来源岗位：{meta.position_name}",
            f"- 成熟度：{meta.maturity}",
            f"- 学习重点：{meta.learn_from}",
            f"- 不建议学习：{meta.do_not_learn}",
            "",
            "| 原始序号 | 行为要项 | 专员行为标准 | 主管行为标准 | 经理行为标准 | 总监行为标准 |",
            "|---|---|---|---|---|---|",
        ]
        for row in items:
            lines.append(
                "| "
                + " | ".join(
                    [
                        md_cell(row["序号"]),
                        md_cell(row["行为要项"]),
                        md_cell(row["专员行为标准"]),
                        md_cell(row["主管行为标准"]),
                        md_cell(row["经理行为标准"]),
                        md_cell(row["总监行为标准"]),
                    ]
                )
                + " |"
            )
        output.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")
        outputs.append(output)
    return outputs


def write_item_and_level_snippets(meta: CaseMeta, rows: list[dict[str, str]]):
    item_outputs = []
    level_outputs = []
    for idx, row in enumerate(rows, 1):
        base_name = f"{meta.case_id:02d}_{filename_part(meta.position_name)}_{idx:02d}_{filename_part(row['行为模块'])}_{filename_part(row['行为要项'])}"
        item_output = BY_BEHAVIOR_ITEM / f"{base_name}.md"
        level_output = BY_LEVEL_PROGRESSION / f"{base_name}.md"
        item_lines = [
            frontmatter(
                meta,
                {
                    "案例类型": "行为要项片段",
                    "行为模块": row["行为模块"],
                    "行为要项": row["行为要项"],
                    "原始序号": row["序号"],
                },
            ),
            "",
            f"# 行为要项片段：{row['行为要项']}",
            "",
            f"- 来源岗位：{meta.position_name}",
            f"- 行为模块：{row['行为模块']}",
            f"- 学习重点：{meta.learn_from}",
            f"- 不建议学习：{meta.do_not_learn}",
            "",
            "| 职级 | 行为标准 |",
            "|---|---|",
        ]
        level_lines = [
            frontmatter(
                meta,
                {
                    "案例类型": "职级递进片段",
                    "行为模块": row["行为模块"],
                    "行为要项": row["行为要项"],
                    "原始序号": row["序号"],
                },
            ),
            "",
            f"# 职级递进片段：{row['行为要项']}",
            "",
            f"- 来源岗位：{meta.position_name}",
            f"- 行为模块：{row['行为模块']}",
            f"- 成熟度：{meta.maturity}",
            f"- 使用提醒：{meta.do_not_learn}",
            "",
            "| 职级 | 行为标准 |",
            "|---|---|",
        ]
        for level in LEVEL_HEADERS:
            level_name = level.replace("行为标准", "")
            standard = md_cell(row[level])
            item_lines.append(f"| {level_name} | {standard} |")
            level_lines.append(f"| {level_name} | {standard} |")
        item_output.write_text("\n".join(item_lines).rstrip() + "\n", encoding="utf-8")
        level_output.write_text("\n".join(level_lines).rstrip() + "\n", encoding="utf-8")
        item_outputs.append(item_output)
        level_outputs.append(level_output)
    return item_outputs, level_outputs


def write_position_map():
    path = SOURCE_DIR / "职位图谱.xlsx"
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    rows = []
    for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row, min_col=1, max_col=4, values_only=True):
        values = [text(value) for value in row]
        if any(values):
            rows.append(values)
    class_counter = Counter(row[1] for row in rows if row[1])
    lines = [
        "---",
        '来源文件: "任职资格参考标准/职位图谱.xlsx"',
        '用途: "校验职类、职务子类、标准岗位名称"',
        "---",
        "",
        "# 职位图谱",
        "",
        "## 职类统计",
        "",
        "| 职类 | 岗位数量 |",
        "|---|---:|",
    ]
    for job_class, count in sorted(class_counter.items()):
        lines.append(f"| {md_cell(job_class)} | {count} |")
    lines.extend(["", "## 明细", "", "| 职族 | 职类 | 职务子类 | 标准岗位 |", "|---|---|---|---|"])
    for row in rows:
        lines.append("| " + " | ".join(md_cell(value) for value in row) + " |")
    POSITION_MAP_OUTPUT.write_text("\n".join(lines).rstrip() + "\n", encoding="utf-8")
    return POSITION_MAP_OUTPUT, len(rows), len(class_counter)


def ensure_dirs():
    for directory in [FULL_POSITIONS, BY_MODULE, BY_BEHAVIOR_ITEM, BY_LEVEL_PROGRESSION, POSITION_MAP_OUTPUT.parent]:
        directory.mkdir(parents=True, exist_ok=True)


def clear_generated_outputs():
    for directory in [FULL_POSITIONS, BY_MODULE, BY_BEHAVIOR_ITEM, BY_LEVEL_PROGRESSION]:
        if directory.exists():
            for path in directory.glob("*.md"):
                path.unlink()
    for path in [POSITION_MAP_OUTPUT, REPORT_OUTPUT]:
        if path.exists():
            path.unlink()


def convert():
    ensure_dirs()
    clear_generated_outputs()
    report_lines = [
        "# 案例库转换报告",
        "",
        "本报告由 `scripts/excel_to_markdown_converter.py` 自动生成。",
        "",
    ]
    total_full = total_modules = total_items = total_levels = 0
    for meta in CASES:
        path = SOURCE_DIR / meta.source_file
        sheet_name, header_row, rows = read_case_rows(path)
        full_output = write_full_position(meta, sheet_name, header_row, rows)
        module_outputs = write_module_snippets(meta, rows)
        item_outputs, level_outputs = write_item_and_level_snippets(meta, rows)
        blank_counts = {level: sum(1 for row in rows if not row[level]) for level in LEVEL_HEADERS}
        org_rows = sum(1 for row in rows if row["行为模块"] == "组织贡献")
        normalized_items = [
            row
            for row in rows
            if row.get("原始行为要项", row["行为要项"]) != row["行为要项"]
        ]
        total_full += 1
        total_modules += len(module_outputs)
        total_items += len(item_outputs)
        total_levels += len(level_outputs)
        report_lines.extend(
            [
                f"## {meta.position_name}",
                "",
                f"- 来源文件：`{path}`",
                f"- 使用 Sheet：{sheet_name}",
                f"- 表头行：{header_row}",
                f"- 完整案例：`{full_output.relative_to(SKILL_ROOT)}`",
                f"- 有效数据行：{len(rows)}",
                f"- 行为模块片段：{len(module_outputs)}",
                f"- 行为要项片段：{len(item_outputs)}",
                f"- 职级递进片段：{len(level_outputs)}",
                f"- 组织贡献行数：{org_rows}",
                "- 空白职级单元格：" + "；".join(f"{level}={count}" for level, count in blank_counts.items()),
                "- 统一调整："
                + (
                    "；".join(
                        f"序号{row['序号']} {row['原始行为要项']} -> {row['行为要项']}"
                        for row in normalized_items
                    )
                    if normalized_items
                    else "无"
                ),
                f"- 成熟度：{meta.maturity}",
                f"- 不建议学习：{meta.do_not_learn}",
                "",
            ]
        )
    position_map, position_count, class_count = write_position_map()
    report_lines.extend(
        [
            "## 汇总",
            "",
            f"- 完整岗位案例：{total_full}",
            f"- 行为模块片段：{total_modules}",
            f"- 行为要项片段：{total_items}",
            f"- 职级递进片段：{total_levels}",
            f"- 职位图谱明细：{position_count} 条，职类 {class_count} 个",
            f"- 职位图谱输出：`{position_map.relative_to(SKILL_ROOT)}`",
            "",
        ]
    )
    REPORT_OUTPUT.write_text("\n".join(report_lines).rstrip() + "\n", encoding="utf-8")
    return REPORT_OUTPUT


if __name__ == "__main__":
    output = convert()
    print(f"转换完成：{output}")
